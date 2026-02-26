"""
XLSX writer para Tese IAMSPE.
Gera planilha pivot: linhas=meses, colunas=rubricas IAMSPE, última=VALOR DEVIDO.
Formato idêntico à planilha de referência do escritório.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def write_iamspe_xlsx(resultado: dict, output_path: str) -> str:
    """
    Gera XLSX no formato de acúmulo IAMSPE.

    Args:
        resultado: dict de TeseIAMSPE.processar()
        output_path: caminho do arquivo de saída

    Returns:
        output_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Acúmulo IAMSPE"

    # --- Estilos ---
    AZUL = "1A365D"
    DOURADO = "C7A76D"
    money_fmt = '#,##0.00'

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    total_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=13, color=AZUL)
    subtitle_font = Font(bold=True, size=11, color=DOURADO)

    rubricas = resultado['rubricas']       # OrderedDict {code: label}
    periodos = resultado['periodos']       # OrderedDict {period: {code: value}}
    sorted_codes = list(rubricas.keys())

    # Cálculo de colunas: DATA + rubricas + VALOR DEVIDO
    total_col = 1 + len(sorted_codes) + 1
    last_col_letter = get_column_letter(total_col)

    # --- Cabeçalho do documento ---
    ws.merge_cells(f'A1:{last_col_letter}1')
    t = ws['A1']
    t.value = "CONTRIBUIÇÃO DO IAMSPE SOBRE O SEGUNDO VÍNCULO"
    t.font = title_font
    t.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col_letter}2')
    s = ws['A2']
    s.value = f"REQUERENTE: {resultado['nome_cliente']}"
    s.font = subtitle_font
    s.alignment = Alignment(horizontal='center')

    # --- Linha de cabeçalho da tabela ---
    HDR = 4

    c = ws.cell(row=HDR, column=1, value="DATA DE PAGAMENTO")
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, code in enumerate(sorted_codes):
        col = 2 + i
        c = ws.cell(row=HDR, column=col, value=rubricas[code])
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    c = ws.cell(row=HDR, column=total_col, value="VALOR DEVIDO:")
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    # --- Linhas de dados ---
    data_start = HDR + 1
    sorted_periods = list(periodos.keys())

    for i, per in enumerate(sorted_periods):
        row = data_start + i
        yyyy, mm = per.split('-')

        # Coluna A: data de pagamento
        c = ws.cell(row=row, column=1, value=f"{mm}/{yyyy}")
        c.border = thin
        c.alignment = Alignment(horizontal='center')

        # Colunas de rubricas
        for j, code in enumerate(sorted_codes):
            col = 2 + j
            val = periodos[per].get(code, 0.0)
            c = ws.cell(row=row, column=col, value=val if val != 0.0 else None)
            c.number_format = money_fmt
            c.border = thin

        # Coluna VALOR DEVIDO = SUM(B{row}:{prev_col}{row})
        data_last_col = get_column_letter(total_col - 1)
        c = ws.cell(row=row, column=total_col)
        c.value = f"=SUM(B{row}:{data_last_col}{row})"
        c.number_format = money_fmt
        c.border = thin

    # --- Linha de totais ---
    last_data = data_start + len(sorted_periods) - 1
    total_row = last_data + 1

    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = total_font
    c.fill = total_fill
    c.border = thin

    for j, code in enumerate(sorted_codes):
        col = 2 + j
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col)
        c.value = f"=SUM({col_letter}{data_start}:{col_letter}{last_data})"
        c.number_format = money_fmt
        c.font = total_font
        c.fill = total_fill
        c.border = thin

    # Total geral (coluna VALOR DEVIDO)
    total_col_letter = get_column_letter(total_col)
    c = ws.cell(row=total_row, column=total_col)
    c.value = f"=SUM({total_col_letter}{data_start}:{total_col_letter}{last_data})"
    c.number_format = money_fmt
    c.font = total_font
    c.fill = total_fill
    c.border = thin

    # --- Largura das colunas ---
    ws.column_dimensions['A'].width = 18
    for j in range(len(sorted_codes)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 24
    ws.column_dimensions[get_column_letter(total_col)].width = 16

    # Altura da linha de cabeçalho (para labels longos)
    ws.row_dimensions[HDR].height = 50

    wb.save(output_path)
    return output_path
