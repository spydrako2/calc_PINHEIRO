"""
Unified XLSX writer for tese results.
Generates auditable spreadsheets with decomposed formulas and atrasado tracking.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment

from src.teses.base_tese import BaseTese


def write_reflexo_xlsx(resultado: dict, output_path: str) -> str:
    """
    Write tese result to XLSX.

    Args:
        resultado: dict from BaseTese.processar()
        output_path: path for output file

    Returns:
        output_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reflexo"

    # --- Styles ---
    AZUL = "1A365D"
    DOURADO = "C7A76D"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    atrasado_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color=AZUL)
    subtitle_font = Font(bold=True, size=11, color=DOURADO)

    # Detect if any period has sexta parte (to decide whether to show those columns)
    periodos_check = resultado['periodos']
    has_sexta = any(d.get('tem_sexta_parte', False) for d in periodos_check.values())

    # Number of columns: 5 base + 3 sexta-parte columns when applicable
    last_col = 'H' if has_sexta else 'E'

    # --- Title block ---
    ws.merge_cells(f'A1:{last_col}1')
    cell_title = ws['A1']
    cell_title.value = f"HoleritePRO — {resultado['tese_nome']}"
    cell_title.font = title_font
    cell_title.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col}2')
    cell_cliente = ws['A2']
    cell_cliente.value = f"Cliente: {resultado['nome_cliente']}"
    cell_cliente.font = subtitle_font
    cell_cliente.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A3:{last_col}3')
    cell_desc = ws['A3']
    cell_desc.value = resultado['tese_descricao']
    cell_desc.font = Font(italic=True, size=9, color="666666")
    cell_desc.alignment = Alignment(horizontal='center', wrap_text=True)

    # --- Headers (row 5) ---
    header_row = 5
    headers = [
        "Competência",
        resultado['verba_nome'],
        "Qtde Quinquênios",
        "% Adic. Temporal",
        "Reflexo Quinquênio",
    ]
    if has_sexta:
        headers += ["Tem 6ª Parte?", "Reflexo 6ª Parte", "TOTAL DEVIDO"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin

    # --- Data rows ---
    periodos = resultado['periodos']
    sorted_periods = sorted(periodos.keys())
    data_start = header_row + 1

    for i, per in enumerate(sorted_periods):
        row = data_start + i
        data = periodos[per]
        yyyy, mm = per.split('-')

        # A: Competência
        ws.cell(row=row, column=1, value=f"{mm}/{yyyy}").border = thin

        # B: Verba — auditable formula with atrasado breakdown
        normal = data['normal']
        atrasados = data['atrasados']

        if atrasados:
            parts = []
            if normal != 0:
                parts.append(f"{normal:.2f}")
            for _, val in atrasados:
                parts.append(f"{val:.2f}")

            formula = "+".join(parts) if parts else "0"
            cell_b = ws.cell(row=row, column=2)
            cell_b.value = f"={formula}"
            cell_b.number_format = money_fmt
            cell_b.border = thin
            cell_b.fill = atrasado_fill

            # Comment with breakdown (pagamento = comp + 1 mês)
            comment_lines = []
            if normal != 0:
                comment_lines.append(f"Normal: R$ {normal:.2f}")
            for comp_pgto, val in atrasados:
                pgto_display = BaseTese.format_comp_display(
                    BaseTese.mes_pagamento(comp_pgto)
                )
                comment_lines.append(f"Atraso pago em {pgto_display}: R$ {val:.2f}")
            cell_b.comment = Comment("\n".join(comment_lines), "HoleritePRO")
        else:
            cell_b = ws.cell(row=row, column=2, value=normal)
            cell_b.number_format = money_fmt
            cell_b.border = thin

        # C: Quinquênios
        cell_c = ws.cell(row=row, column=3, value=data['quinquenios'])
        cell_c.border = thin
        cell_c.alignment = Alignment(horizontal='center')

        # D: % = C * 5%
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f"=C{row}*5%"
        cell_d.number_format = pct_fmt
        cell_d.border = thin

        # E: Reflexo = B * D
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f"=B{row}*D{row}"
        cell_e.number_format = money_fmt
        cell_e.border = thin

        if has_sexta:
            # F: Tem 6ª Parte?
            sexta_val = "Sim" if data.get('tem_sexta_parte', False) else "Não"
            cell_f = ws.cell(row=row, column=6, value=sexta_val)
            cell_f.border = thin
            cell_f.alignment = Alignment(horizontal='center')

            # G: Reflexo 6ª Parte = IF(F="Sim", E/6, 0)
            cell_g = ws.cell(row=row, column=7)
            cell_g.value = f'=IF(F{row}="Sim",E{row}/6,0)'
            cell_g.number_format = money_fmt
            cell_g.border = thin

            # H: Total Devido = E + G
            cell_h = ws.cell(row=row, column=8)
            cell_h.value = f"=E{row}+G{row}"
            cell_h.number_format = money_fmt
            cell_h.border = thin

    # --- Totals row ---
    last_data = data_start + len(sorted_periods) - 1
    total_row = last_data + 1

    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).border = thin

    cell_tb = ws.cell(row=total_row, column=2)
    cell_tb.value = f"=SUM(B{data_start}:B{last_data})"
    cell_tb.number_format = money_fmt
    cell_tb.font = total_font
    cell_tb.border = thin

    cell_te = ws.cell(row=total_row, column=5)
    cell_te.value = f"=SUM(E{data_start}:E{last_data})"
    cell_te.number_format = money_fmt
    cell_te.font = total_font
    cell_te.border = thin

    if has_sexta:
        cell_tg = ws.cell(row=total_row, column=7)
        cell_tg.value = f"=SUM(G{data_start}:G{last_data})"
        cell_tg.number_format = money_fmt
        cell_tg.font = total_font
        cell_tg.border = thin

        cell_th = ws.cell(row=total_row, column=8)
        cell_th.value = f"=SUM(H{data_start}:H{last_data})"
        cell_th.number_format = money_fmt
        cell_th.font = total_font
        cell_th.border = thin

    # --- Legend ---
    legend_row = total_row + 2
    ws.cell(row=legend_row, column=1, value="Legenda:").font = Font(bold=True, size=9)
    ws.cell(row=legend_row + 1, column=1, value="Células amarelas").fill = atrasado_fill
    ws.cell(row=legend_row + 1, column=1).font = Font(size=9)
    ws.cell(row=legend_row + 1, column=2,
            value="Contêm valores com atraso consolidados. Passe o mouse para ver detalhamento.").font = Font(size=9)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    if has_sexta:
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18

    wb.save(output_path)
    return output_path
