"""
XLSX writer para Tese APEOESP — Quinquênio e Sexta Parte sobre Gratificações.
Estrutura por ano: 12 meses + linha 13° Salário + linha 1/3 Férias.
Formato idêntico à planilha de referência do escritório.
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from src.teses.base_tese import BaseTese


# Colunas fixas (índices 1-based)
COL_DATA     = 1
COL_GRATIF   = 2
COL_GTE      = 3
COL_GAM      = 4
COL_TOTAL_V  = 5
COL_QUINQ    = 6
COL_PCT      = 7
COL_DQUINQ   = 8
COL_SEXTA    = 9
COL_D6P      = 10
COL_TOTAL    = 11

HEADERS = [
    "DATA DE PAGAMENTO",
    "GRATIFICAÇÃO GERAL",
    "GTE - GRATIFICAÇÃO POR TRABALHO EDUCACIONAL",
    "GAM - GRATIFICAÇÃO POR ATIVIDADE DE MAGISTÉRIO",
    "TOTAL VANTAGENS INTEGRAIS",
    "QTDE. QUINQUÊNIOS",
    "PORCENTAGEM",
    "DIFERENÇA QUINQUÊNIOS",
    "TEM 6ª PARTE?",
    "DIFERENÇA 6ª PARTE",
    "TOTAL DEVIDO",
]


def _col_formula(normal: float, atrasados: list):
    """Returns formula string '=normal+atraso1+...' or plain value when no atrasados."""
    if not atrasados:
        return normal or None
    parts = []
    if normal:
        parts.append(f"{normal:.2f}")
    for _, val in atrasados:
        parts.append(f"{val:.2f}")
    return ("=" + "+".join(parts)) if parts else None


def write_apeoesp_xlsx(resultado: dict, output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quinquênio e 6ª Parte"

    # --- Estilos ---
    AZUL   = "1A365D"
    DOURADO = "C7A76D"
    VERDE  = "C6EFCE"   # 13° e 1/3 férias
    CINZA  = "E2E8F0"   # total

    money_fmt = '#,##0.00'
    pct_fmt   = '0.00%'

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    hdr_font   = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill   = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
    total_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=13, color=AZUL)
    sub_font   = Font(bold=True, size=11, color=DOURADO)
    verde_fill    = PatternFill(start_color=VERDE,    end_color=VERDE,    fill_type="solid")
    cinza_fill    = PatternFill(start_color=CINZA,    end_color=CINZA,    fill_type="solid")
    atrasado_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold9      = Font(bold=True, size=9)
    norm9      = Font(size=9)

    last_col = get_column_letter(COL_TOTAL)

    # --- Cabeçalho do documento ---
    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value = "QUINQUÊNIO E SEXTA PARTE — GRATIFICAÇÕES APEOESP"
    t.font = title_font
    t.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col}2')
    s = ws['A2']
    s.value = f"REQUERENTE: {resultado['nome_cliente']}"
    s.font = sub_font
    s.alignment = Alignment(horizontal='center')

    # --- Linha de cabeçalho da tabela ---
    HDR = 4
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=HDR, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # --- Agrupar períodos por ano ---
    periodos = resultado['periodos']
    years = defaultdict(list)
    for per in sorted(periodos.keys()):
        years[per[:4]].append(per)

    current_row = HDR + 1

    # Rastrear linhas de TOTAL_DEVIDO para o total final
    all_total_rows = []   # linhas de meses normais
    all_13_rows    = []   # linhas de 13° por ano
    all_13f_rows   = []   # linhas de 1/3 férias por ano

    for year, periods_in_year in sorted(years.items()):
        year_data_start = current_row
        year_total_rows = []

        for per in periods_in_year:
            row = current_row
            d = periodos[per]
            yyyy, mm = per.split('-')

            def w(col, val=None, fmt=None, bold=False, fill=None, center=False):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin
                if fmt:   c.number_format = fmt
                if bold:  c.font = bold9
                if fill:  c.fill = fill
                if center: c.alignment = Alignment(horizontal='center')
                else:      c.font = norm9
                return c

            # A: Data
            c = ws.cell(row=row, column=COL_DATA, value=f"{mm}/{yyyy}")
            c.border = thin
            c.font = norm9
            c.alignment = Alignment(horizontal='center')

            # B: Gratif Geral — formula if has atrasados
            _gn = d.get('gratif_geral_normal', d.get('gratif_geral', 0.0))
            _ga = d.get('gratif_geral_atrasados', [])
            c = ws.cell(row=row, column=COL_GRATIF, value=_col_formula(_gn, _ga))
            c.border = thin
            c.number_format = money_fmt
            c.font = norm9
            if _ga:
                c.fill = atrasado_fill
                comment_lines = []
                if _gn:
                    comment_lines.append(f"Normal: R$ {_gn:.2f}")
                for comp_pgto, val in _ga:
                    pgto_display = BaseTese.format_comp_display(BaseTese.mes_pagamento(comp_pgto))
                    comment_lines.append(f"Atraso pago em {pgto_display}: R$ {val:.2f}")
                c.comment = Comment("\n".join(comment_lines), "HoleritePRO")

            # C: GTE — formula if has atrasados
            _gn = d.get('gte_normal', d.get('gte', 0.0))
            _ga = d.get('gte_atrasados', [])
            c = ws.cell(row=row, column=COL_GTE, value=_col_formula(_gn, _ga))
            c.border = thin
            c.number_format = money_fmt
            c.font = norm9
            if _ga:
                c.fill = atrasado_fill
                comment_lines = []
                if _gn:
                    comment_lines.append(f"Normal: R$ {_gn:.2f}")
                for comp_pgto, val in _ga:
                    pgto_display = BaseTese.format_comp_display(BaseTese.mes_pagamento(comp_pgto))
                    comment_lines.append(f"Atraso pago em {pgto_display}: R$ {val:.2f}")
                c.comment = Comment("\n".join(comment_lines), "HoleritePRO")

            # D: GAM — formula if has atrasados
            _gn = d.get('gam_normal', d.get('gam', 0.0))
            _ga = d.get('gam_atrasados', [])
            c = ws.cell(row=row, column=COL_GAM, value=_col_formula(_gn, _ga))
            c.border = thin
            c.number_format = money_fmt
            c.font = norm9
            if _ga:
                c.fill = atrasado_fill
                comment_lines = []
                if _gn:
                    comment_lines.append(f"Normal: R$ {_gn:.2f}")
                for comp_pgto, val in _ga:
                    pgto_display = BaseTese.format_comp_display(BaseTese.mes_pagamento(comp_pgto))
                    comment_lines.append(f"Atraso pago em {pgto_display}: R$ {val:.2f}")
                c.comment = Comment("\n".join(comment_lines), "HoleritePRO")

            # E: Total Vantagens = SUM(B:D)
            c = ws.cell(row=row, column=COL_TOTAL_V)
            c.value = f"=SUM(B{row}:D{row})"
            c.number_format = money_fmt
            c.border = thin
            c.font = norm9

            # F: Quinquênios
            c = ws.cell(row=row, column=COL_QUINQ, value=d['quinquenios'] or None)
            c.border = thin
            c.font = norm9
            c.alignment = Alignment(horizontal='center')

            # G: Porcentagem = F * 5%
            c = ws.cell(row=row, column=COL_PCT)
            c.value = f"=F{row}*5%"
            c.number_format = pct_fmt
            c.border = thin
            c.font = norm9

            # H: Diferença Quinquênios = E * G
            c = ws.cell(row=row, column=COL_DQUINQ)
            c.value = f"=E{row}*G{row}"
            c.number_format = money_fmt
            c.border = thin
            c.font = norm9

            # I: Tem 6ª Parte?
            sexta_val = "Sim" if d['tem_sexta_parte'] else "Não"
            c = ws.cell(row=row, column=COL_SEXTA, value=sexta_val)
            c.border = thin
            c.font = norm9
            c.alignment = Alignment(horizontal='center')

            # J: Diferença 6ª Parte = IF(I="Sim", H/6, 0)
            c = ws.cell(row=row, column=COL_D6P)
            c.value = f'=IF(I{row}="Sim",(H{row})/6,0)'
            c.number_format = money_fmt
            c.border = thin
            c.font = norm9

            # K: Total Devido = H + J
            c = ws.cell(row=row, column=COL_TOTAL)
            c.value = f"=H{row}+J{row}"
            c.number_format = money_fmt
            c.border = thin
            c.font = norm9

            year_total_rows.append(row)
            all_total_rows.append(row)
            current_row += 1

        # --- 13° Salário (ao final do ano) ---
        row_13 = current_row
        c = ws.cell(row=row_13, column=COL_DATA, value="13 SALÁRIO")
        c.font = bold9
        c.fill = verde_fill
        c.border = thin

        for col in range(2, COL_TOTAL + 1):
            c = ws.cell(row=row_13, column=col)
            c.fill = verde_fill
            c.border = thin
            c.font = norm9

        # K: 13° = SUM(K_meses_do_ano) / 12
        sum_range_k = "+".join(f"K{r}" for r in year_total_rows)
        c = ws.cell(row=row_13, column=COL_TOTAL)
        c.value = f"=({sum_range_k})/12"
        c.number_format = money_fmt
        c.font = bold9
        c.fill = verde_fill
        c.border = thin

        all_13_rows.append(row_13)
        current_row += 1

        # --- 1/3 Férias ---
        row_ferias = current_row
        c = ws.cell(row=row_ferias, column=COL_DATA, value="1/3 FÉRIAS")
        c.font = bold9
        c.fill = verde_fill
        c.border = thin

        for col in range(2, COL_TOTAL + 1):
            c = ws.cell(row=row_ferias, column=col)
            c.fill = verde_fill
            c.border = thin
            c.font = norm9

        # K: 1/3 férias = K_13° / 3
        c = ws.cell(row=row_ferias, column=COL_TOTAL)
        c.value = f"=K{row_13}/3"
        c.number_format = money_fmt
        c.font = bold9
        c.fill = verde_fill
        c.border = thin

        all_13f_rows.append(row_ferias)
        current_row += 1

    # --- Linha de TOTAL GERAL ---
    total_row = current_row + 1
    c = ws.cell(row=total_row, column=COL_DATA, value="TOTAL GERAL")
    c.font = total_font
    c.fill = cinza_fill
    c.border = thin

    for col in range(2, COL_TOTAL + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = cinza_fill
        c.border = thin

    # Soma de todos os meses + 13° + 1/3 férias
    all_k_refs = [f"K{r}" for r in all_total_rows + all_13_rows + all_13f_rows]
    c = ws.cell(row=total_row, column=COL_TOTAL)
    c.value = "=" + "+".join(all_k_refs)
    c.number_format = money_fmt
    c.font = total_font
    c.fill = cinza_fill
    c.border = thin

    # --- Largura das colunas ---
    widths = [16, 20, 34, 34, 20, 16, 14, 22, 14, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Altura do cabeçalho
    ws.row_dimensions[HDR].height = 55

    wb.save(output_path)
    return output_path
